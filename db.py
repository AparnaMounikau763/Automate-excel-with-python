import openpyxl

# Open the workbook
wb = openpyxl.load_workbook('excel.1.xlsx')

# Get the active sheet
ws = wb.active

# Add some data to the sheet
ws['A1'] = 'Name'
ws['B1'] = 'Roll Number'
ws['C1'] = 'Section'
ws['A2'] = 'Puhspa'
ws['B2'] = 4389
ws['C2'] = 'CAI'
ws['A3'] = 'Anjali'
ws['B3'] = 4364
ws['C3'] = 'CAI'
ws['A4'] = 'Yogitha'
ws['B4'] = 4392
ws['C4'] = 'CAI'
ws['A5'] = 'Aparna'
ws['B5'] = "43B8"
ws['C5'] = 'CAI'
ws['A6'] = 'siri'
ws['B6'] = '43E9'
ws['C6'] = 'AID'
ws['A7'] = 'Prabha'
ws['B7'] = '45C8'
ws['C7'] = 'AID'

# Save the workbook
wb.save('excel.1.xlsx')